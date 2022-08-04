import psutil
import subprocess
import openpyxl
import datetime
import schedule
import time

time_interval = int(input("Imput time interval"))
path_to_save = input("Where to save: ")
subprocess.call(input("Path to executable : "))


def monitor():
    p = psutil.Process()
    open_handles = p.num_handles()
    time_to_insert = datetime.datetime.now().strftime("%Y%m%d-%h:%M:%S")
    cpu = p.cpu_percent(interval=1)
    memory_mb = p.memory_percent()
    file = openpyxl.load_workbook(path_to_save)
    sheet = file.active
    if sheet.cell(column=1, row=1).value is None:
        sheet.cell(column=1, row=1, value='time')
        sheet.cell(column=2, row=1, value='open hendles')
        sheet.cell(column=3, row=1, value='cpu')
        sheet.cell(column=5, row=1, value='memory mb')
    sheet.cell(column=1, row=sheet.max_row + 1, value=open_handles)
    sheet.cell(column=2, row=sheet.max_row, value=time_to_insert)
    sheet.cell(column=3, row=sheet.max_row, value=cpu)
    sheet.cell(column=5, row=sheet.max_row, value=memory_mb)
    file.save(path_to_save)


schedule.every(time_interval).do(monitor())
while True:
    schedule.run_pending()
    time.sleep(1)
